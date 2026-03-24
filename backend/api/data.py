from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import io, openpyxl
from faker import Faker

from core.database import get_db
from core.auth import get_current_user
from models.user import User

router = APIRouter(prefix="/api/data", tags=["data"])
fake = Faker()


class VariableConfig(BaseModel):
    names: List[str]


class ParsedExcelResponse(BaseModel):
    headers: List[str]
    rows: List[Dict[str, Any]]
    total_rows: int
    errors: List[str]


class DummyDataRequest(BaseModel):
    variable_names: List[str]
    count: int = 10


@router.post("/parse-excel", response_model=ParsedExcelResponse)
async def parse_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload and parse Excel/CSV file, returning preview data."""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) and CSV files supported")

    content = await file.read()
    errors = []
    rows = []
    headers = []

    try:
        if file.filename.endswith('.csv'):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
            headers = [h.strip() for h in (reader.fieldnames or [])]
            rows = []
            for row in reader:
                rows.append({h.strip(): str(v).strip() for h, v in row.items()})
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return ParsedExcelResponse(headers=[], rows=[], total_rows=0, errors=["Empty file"])
            headers = [str(h).strip() for h in all_rows[0] if h is not None]
            for r in all_rows[1:]:
                row_dict = {}
                for i, h in enumerate(headers):
                    val = str(r[i]).strip() if i < len(r) and r[i] is not None else ""
                    row_dict[h] = val
                rows.append(row_dict)
            wb.close()

        # Validate email column
        email_cols = [h for h in headers if 'receivers email' in h.lower() or 'email' in h.lower()]
        if not email_cols:
            errors.append("Warning: No 'Receivers Email' column found. Please ensure one column is named 'Receivers Email'.")

        # Validate email values
        import re
        email_col = email_cols[0] if email_cols else None
        if email_col:
            for i, row in enumerate(rows, start=2):
                raw_emails = row.get(email_col, "").strip()
                if not raw_emails:
                    errors.append(f"Row {i}: Empty email")
                else:
                    # Split by common separators and validate each
                    parts = [p.strip() for p in re.split(r'[,;|\n]+', raw_emails) if p.strip()]
                    for p in parts:
                        if not re.match(r'^[^@]+@[^@]+\.[^@]+$', p):
                            errors.append(f"Row {i}: Invalid email '{p}'")

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return ParsedExcelResponse(
        headers=headers,
        rows=rows[:100],  # Preview max 100 rows
        total_rows=len(rows),
        errors=errors,
    )


@router.get("/sample-excel")
async def download_sample_excel(
    variables: str = "name,company,position",  # comma-separated
    current_user: User = Depends(get_current_user),
):
    """Generate and download sample Excel with given variable columns."""
    var_list = [v.strip() for v in variables.split(",") if v.strip()]
    headers = ["Receivers Email"] + var_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    # Add 3 sample rows
    samples = [
        {"Receivers Email": "john.doe@example.com", "name": "John Doe", "company": "Acme Corp", "position": "Manager"},
        {"Receivers Email": "jane.smith@example.com, tech@ltd.com", "name": "Jane Smith", "company": "Tech Ltd", "position": "Director"},
        {"Receivers Email": "bob.jones@example.com", "name": "Bob Jones", "company": "Global Inc", "position": "CEO"},
    ]
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(header, f"sample_{header}"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sample_contacts.xlsx"}
    )


@router.post("/generate-dummy")
async def generate_dummy_data(
    data: DummyDataRequest,
    current_user: User = Depends(get_current_user),
):
    """Generate fake data based on variable names."""
    FAKER_MAP = {
        "name": lambda: fake.name(),
        "first_name": lambda: fake.first_name(),
        "last_name": lambda: fake.last_name(),
        "company": lambda: fake.company(),
        "position": lambda: fake.job(),
        "phone": lambda: fake.phone_number(),
        "city": lambda: fake.city(),
        "country": lambda: fake.country(),
        "address": lambda: fake.address(),
        "website": lambda: fake.url(),
        "industry": lambda: fake.bs().split()[0].title(),
    }

    rows = []
    for _ in range(min(data.count, 100)):
        row = {"Receivers Email": fake.email()}
        for var in data.variable_names:
            key = var.lower().strip()
            row[var] = FAKER_MAP.get(key, lambda: fake.word())()
        rows.append(row)

    return {"rows": rows, "count": len(rows)}
